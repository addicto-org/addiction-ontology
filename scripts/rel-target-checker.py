import os
import openpyxl
from collections import Counter
import argparse
import sys
import requests
import pyhornedowl
from urllib.request import urlopen
import csv

PREFIXES = [ ["ADDICTO","http://addictovocab.org/ADDICTO_"],
             ["BFO","http://purl.obolibrary.org/obo/BFO_"],
             ["CHEBI","http://purl.obolibrary.org/obo/CHEBI_"],
             ["UBERON","http://purl.obolibrary.org/obo/UBERON_"],
             ["PATO","http://purl.obolibrary.org/obo/PATO_"],
             ["BCIO","http://humanbehaviourchange.org/ontology/BCIO_"],
             ["SEPIO","http://purl.obolibrary.org/obo/SEPIO_"],
             ["OMRSE","http://purl.obolibrary.org/obo/OMRSE_"],
             ["OBCS","http://purl.obolibrary.org/obo/OBCS_"],
             ["OGMS","http://purl.obolibrary.org/obo/OGMS_"],
             ["ENVO","http://purl.obolibrary.org/obo/ENVO_"],
             ["OBI", "http://purl.obolibrary.org/obo/OBI_"],
             ["MFOEM","http://purl.obolibrary.org/obo/MFOEM_"],
             ["MF","http://purl.obolibrary.org/obo/MF_"],
             ["CHMO","http://purl.obolibrary.org/obo/CHMO_"],
             ["DOID","http://purl.obolibrary.org/obo/DOID_"],
             ["IAO","http://purl.obolibrary.org/obo/IAO_"],
             ["ERO","http://purl.obolibrary.org/obo/ERO_"],
             ["PO","http://purl.obolibrary.org/obo/PO_"],
             ["RO","http://purl.obolibrary.org/obo/RO_"],
             ["APOLLO_SV","http://purl.obolibrary.org/obo/APOLLO_SV_"],
             ["PDRO","http://purl.obolibrary.org/obo/PDRO_"],
             ["GAZ","http://purl.obolibrary.org/obo/GAZ_"],
             ["GSSO","http://purl.obolibrary.org/obo/GSSO_"]
           ]

# recursive getListOfFiles from https://thispointer.com/python-how-to-get-list-of-files-in-directory-and-sub-directories/
def getListOfFiles(dirName):
    # create a list of file and sub directories 
    # names in the given directory 
    listOfFile = os.listdir(dirName)
    allFiles = list()
    # Iterate over all the entries
    for entry in listOfFile:
        # Create full path
        fullPath = os.path.join(dirName, entry)
        # If entry is a directory then get the list of files in this directory 
        if os.path.isdir(fullPath):
            allFiles = allFiles + getListOfFiles(fullPath)
        else:
            allFiles.append(fullPath)
                
    return allFiles    

if __name__ == '__main__':

    parser=argparse.ArgumentParser()
    parser.add_argument('--path', '-i',help='Name of the path')
    parser.add_argument('--name', '-n', help='Name of ontology')
    
    args=parser.parse_args()
    path = args.path + "/"
    ontology_name = args.name
    
    
    if args is None :
        parser.print_help()
        sys.exit('Not enough arguments. Expected at least -i "Path to ontology folder" "Type of ontology" ')

    
    # print("ontology_name is: ", ontology_name)
    id_list = []
    label_list = []
    # path = "/home/tom/Documents/PROGRAMMING/Python/addiction-ontology/inputs" #test
    files = getListOfFiles(path) 
    # print("files are: ", files)
    

    allInfo = []
    if ontology_name == "ADDICTO":
        #todo: is there a better way than below? 
        inputFileName = path.rsplit("/", 2)[0] + "/imports/External_Imports_New_Labels.csv"
        # inputFileName = "/home/tom/Documents/PROGRAMMING/Python/addiction-ontology/imports/External_Imports_New_Labels.csv" #test
        with open (inputFileName, newline='') as csvfile:
            data = csv.reader(csvfile, delimiter=',', quotechar='"')
            count = 0
            for row in data:
                count = count+1
                if count > 2:
                    values = {}
                    if row[0] != None and row[0] != "":            
                        values["ID"] = row[0]
                    if row[1] != None and row[1].strip() != "":
                        values["Label"] = row[1]   
                    if values:   
                        if "Label" in values:   
                            if values["Label"] != None and values["Label"].strip() != "":          
                                allInfo.append(values) # all ID, Label 
        # print("allInfo for External_Imports_New_Labels.csv is: ", allInfo)

    id_list = []
    label_list = []
    for f in files:
        inputFileName = f
        # print(inputFileName)
        #error checking:
        fileName, file_extension = os.path.splitext(inputFileName)
        if file_extension == ".xlsx":
            wb = openpyxl.load_workbook(inputFileName) 
            sheet = wb.active
            data = sheet.rows
            rows = []
            header = [i.value for i in next(data)]
            # print("header is: ", header)
            for row in sheet[2:sheet.max_row]:
                values = {}
                for key, cell in zip(header, row):
                    if key == "ID" and cell.value != None and cell.value.strip() != "":
                        id_list.append(cell.value) # just the ID's
                        values[key] = cell.value
                    if key == "Label" and cell.value != None and cell.value.strip() != "":
                        values[key] = cell.value
                if values:
                    if "Label" in values:
                        if values["Label"] != None and values["Label"].strip() != "":
                            allInfo.append(values) # all ID, Label 
# print("allInfo is: ", allInfo)

if ontology_name == "ADDICTO":
    location = f"https://raw.githubusercontent.com/addicto-org/addiction-ontology/master/addicto_external.owx"
else: 
    #todo: add BCIO ontology below - do we need to support more?:
    location = f"https://raw.githubusercontent.com/HumanBehaviourChangeProject/ontologies/master/Upper%20Level%20BCIO/bcio-merged.owx"
# get ID's and Labels using pyhornedowl from above .owx:
data = urlopen(location).read()  # bytes
ontofile1 = data.decode('utf-8')

ontology1 = pyhornedowl.open_ontology(ontofile1)

for prefix in PREFIXES:
    ontology1.add_prefix_mapping(prefix[0], prefix[1])

for termid in ontology1.get_classes():
    termshortid = ontology1.get_id_for_iri(termid)
    RDFSLABEL = "http://www.w3.org/2000/01/rdf-schema#label"
    label = ontology1.get_annotation(termid, RDFSLABEL)
    if label != None and label.strip() != "":
        # DEFINITION = "http://purl.obolibrary.org/obo/IAO_0000115"
        # definition = ontology1.get_annotation(termid, DEFINITION)
        term_entry = {'ID': termid if termshortid is None else termshortid,
                        'Label': label.strip(),
                        # 'definition': definition}
                    }
        
        allInfo.append(term_entry) # join up with allInfo
# with open('allInfoBCIO.txt', 'a') as f:
#     f.write(str(allInfo))
# print(allInfo)
#todo: does the combined allInfo list contain duplicates? Does it matter?

# check against ID's and Rel columns in all spreadsheets:

missingParents = []

for f in files:
        inputFileName = f
        fileName, file_extension = os.path.splitext(inputFileName)
        if file_extension == ".xlsx":
            wb = openpyxl.load_workbook(inputFileName) 
            sheet = wb.active
            data = sheet.rows
            rows = []
            header = [i.value for i in next(data)]

            for row in sheet[2:sheet.max_row]:
                values = {}
                for key, cell in zip(header, row):
                    if key == "Label":
                        label1 = cell.value 
                    if key == "ID" and cell.value != None:
                        #using ontology_name to support "BCIO" as well here..
                        if ontology_name in cell.value: #don't check parent where ID is external
                            for key2, cell2 in zip(header, row):
                                if key2 == "Parent":
                                    if cell2.value != None and cell2.value != '':
                                        match = False
                                        for item in allInfo:                                
                                            for key3, value3 in item.items():
                                                if key3 == "Label":
                                                    if value3.strip() == cell2.value.strip():                                            
                                                        match = True
                                        if not match:    
                                            values["Sheet"] = f.rsplit("/", 1)[1]                                
                                            values["ID"] = cell.value
                                            values["Label"] = label1
                                            values["Description"] = "Parent \"" +  cell2.value + "\" does not exist" 
                                            missingParents.append(values) 


#check "REL " columns in all sheets now and append problems to missingParents

for f in files:
        inputFileName = f
        fileName, file_extension = os.path.splitext(inputFileName)
        if file_extension == ".xlsx":
            wb = openpyxl.load_workbook(inputFileName) 
            sheet = wb.active
            data = sheet.rows
            rows = []
            header = [i.value for i in next(data)]

            for row in sheet[2:sheet.max_row]:
                values = {}
                for key, cell in zip(header, row):
                    if key == "Label": 
                        label1 = cell.value 
                    if key == "ID" and cell.value != None:
                        for key2, cell2 in zip(header, row):
                            if "REL" in key2:
                                if cell2.value != None and cell2.value != '':
                                    for i in cell2.value.split(";"): 
                                        match = False
                                        for item in allInfo:                                
                                            for key3, value3 in item.items():
                                                if key3 == "Label":
                                                    if value3.strip() == i.strip():                                            
                                                        match = True
                                        if not match:    
                                            values["Sheet"] = f.rsplit("/", 1)[1]                               
                                            values["ID"] = cell.value
                                            values["Label"] = label1
                                            values["Description"] = "REL target \"" + i + "\" does not exist"
                                            missingParents.append(values)
                                        
count = 0;
for r in missingParents:
        count = count + 1
        print(count, ": ", str(r).strip("{}")) 
        print("")   

#test:
# print(allInfo)                                    
                                        
