import os
import re
import openpyxl
import csv
import pronto
import distutils
import distutils.util

os.chdir("/Users/hastingj/Work/Onto/addiction-ontology/")


def getIdForLabel(value):
    if value in label_id_map.keys():
        return ( label_id_map[value] )
    if value in label_id_map.values():
        return value  # already an ID
    if value.lower() in label_id_map.keys():
        return ( label_id_map[value.lower()] )
    if value.strip() in label_id_map.keys():
        return ( label_id_map[value.strip()] )
    if value.lower().strip() in label_id_map.keys():
        return ( label_id_map[value.lower().strip()] )
    raise ValueError (f"No ID found for {value}.")

def getLabelForID(value):
    if value in label_id_map.values():
        keys = [k for k,v in label_id_map.items() if v == value]
        return keys[0]
    else:
        return value

def getCorrectFormForLabel(value):
    if value in label_id_map.keys():
        return ( value )
    if value.lower() in label_id_map.keys():
        return ( value.lower() )
    if value.strip() in label_id_map.keys():
        return ( value.strip() )
    if value.lower().strip() in label_id_map.keys():
        return ( value.lower().strip() )

    return (value)


# Run the main processing:


in_path = 'inputs'
out_path = 'outputs'
os.makedirs(out_path,exist_ok=True)   # shouldn't exist, just for testing
pattern = 'AddictO(.*).xlsx'
addicto_files = []
CONVERT_TO_LOWERCASE = False

for root, dirs_list, files_list in os.walk(in_path):
    for file_name in files_list:
        if re.match(pattern, file_name):
            full_file_name = os.path.join(root, file_name)
            addicto_files.append(full_file_name)

addicto_files.remove('inputs/AddictO_Upper level.xlsx')
addicto_files.remove('inputs/AddictO_Not_yet_classified_Defs.xlsx')


next_id = 100
digit_count = 7
total_good = 0
label_id_map = {}
external_entities = {}
num_entities = 0
num_ecigo = 0
num_good_ecigo = 0

# First get the accurate next_id based on the input files

for file in addicto_files:
    try:
        wb = openpyxl.load_workbook(file)
    except Exception as e:
        print(e)
        raise Exception("Error! Not able to parse file: "+file)

    sheet = wb.active
    data = sheet.rows

    header = [i.value for i in next(data)]
    if 'E-CigO' in header:
        eCigOCol = header.index('E-CigO')
        skip_ecig = False
    else:
        skip_ecig = True

    for row in data:
        rowdata = [i.value for i in row]
        num_entities = num_entities + 1
        if not skip_ecig:
            ecigo = rowdata[eCigOCol]
            if ecigo:
                try:
                    ecigo = int(ecigo)
                except:
                    try:
                        ecigo = distutils.util.strtobool(ecigo)
                    except:
                        print("Problem parsing value ",ecigo,"in sheet",sheet)
                        ecigo = True
                if ecigo and int(ecigo) == 1:
                    num_ecigo = num_ecigo + 1

        if rowdata[0] and "ADDICTO:" in rowdata[0]:
            idStr = rowdata[0]
            idStr = idStr[(idStr.rindex(':')+1):]
            id = int(idStr)
            if id >= next_id:
                next_id = id+1

print(next_id)

# Now do the main parsing

for file in addicto_files:
    try:
        wb = openpyxl.load_workbook(file)
    except Exception as e:
        print(e)
        raise Exception("Error! Not able to parse file: "+file)

    sheet = wb.active
    data = sheet.rows
    good_entities = {}

    rel_columns = {}

    header = [i.value for i in next(data)]
    if 'E-CigO' in header:
        eCigOCol = header.index('E-CigO')
        skip_ecig = False
    else:
        skip_ecig = True

    label_column = [i for (i,j) in zip(range(len(header)),header) if j=='Label'][0]
    def_column = [i for (i,j) in zip(range(len(header)),header) if j=='Definition'][0]
    parent_column = [i for (i,j) in zip(range(len(header)),header) if j=='Parent'][0]
    if "REL 'has role'" in header:
        rel_columns['has role'] = [i for (i,j) in zip(range(len(header)),header) if j=="REL 'has role'"][0]
    if "REL 'has part'" in header:
        rel_columns['has part'] = [i for (i,j) in zip(range(len(header)),header) if j=="REL 'has part'"][0]

    for row in data:
        rowdata = [i.value for i in row]

        label = rowdata[label_column]
        if label:
            label = label.strip()
        defn = rowdata[def_column]
        parent = rowdata[parent_column]
        if parent:
            parent = parent.strip()

        if label and defn and parent:
            if len(label)>0 and len(defn)>0 and len(parent)>0:
                if label in good_entities.keys():
                    print(f"Appears to be a duplicate label: {label} in {file}")
                good_entities[label] = rowdata

                if not skip_ecig:
                    ecigo = rowdata[eCigOCol]
                    if ecigo:
                        try:
                            ecigo = int(ecigo)
                        except:
                            try:
                                ecigo = distutils.util.strtobool(ecigo)
                            except:
                                print("Problem parsing value ",ecigo,"in sheet",sheet)
                                ecigo = True
                        if ecigo and int(ecigo) == 1:
                            num_good_ecigo = num_good_ecigo + 1

    # Not validating the parents strictly for now.
    print(f"In file {file}, {len(good_entities)} GOOD.")
    total_good = total_good + len(good_entities)

    # Assign them IDs DO THIS IN APP NOW
    for (label,rowdata) in good_entities.items():
        if not rowdata[0] or len(rowdata[0])==0:
            rowdata[0] = "ADDICTO:"+str(next_id).zfill(digit_count)
            next_id = next_id + 1
        elif "ADDICTO:" in rowdata[0]:
            pass
            #print(f"Found existing internal id {rowdata[0]}")
        else:
            #print(f"Found external id {rowdata[0]}")
            external_entities[rowdata[0]] = rowdata
        label_id_map[label] = rowdata[0]

    # If the parent is good too, replace its label with its ID
    for (label,rowdata) in good_entities.items():
        parent = rowdata[parent_column]
        if parent in label_id_map.keys():
            rowdata[parent_column] = label_id_map[parent]
    # Same for relation columns
        for colname, colindex in rel_columns.items():
            if rowdata[colindex]:
                vals = rowdata[colindex].split(";")
                vals_updated = []
                for v in vals:
                    if v in label_id_map.keys():
                        vals_updated.append(label_id_map[v])
                    else:
                        vals_updated.append(v)
                rowdata[colindex] = ";".join(vals_updated)

    # Write to modified files for sending to database
    filename = file.replace(in_path,out_path).split(sep=".")[0] + ".csv"

    with open(filename, 'w') as file:
        writer = csv.writer(file)
        writer.writerow(header)
        for rowdata in good_entities.values():
            writer.writerow(rowdata)

with open('imports/external-entities.csv','w') as file:
    writer = csv.writer(file)
    for rowdata in external_entities.values():
        writer.writerow(rowdata)

print(f"Finished extracting {total_good} good entities.")



if True:
    # # # -----------------------------
    # # # Only needed if IDs were added
    # Write newly generated IDs back to original files ID columns
    for file in addicto_files:

        CHANGED = False
        try:
            wb = openpyxl.load_workbook(file)
        except Exception as e:
            print(e)
            raise Exception("Error! Not able to parse file: "+file)

        sheet = wb.active
        data = sheet.rows

        header = [i.value for i in next(data)]

        label_column = [i for (i,j) in zip(range(len(header)),header) if j=='Label'][0]
        parent_column = [i for (i,j) in zip(range(len(header)),header) if j=='Parent'][0]

        for row in data:
            label = row[label_column].value
            if CONVERT_TO_LOWERCASE and label:
                row[label_column].value = label[0].lower() + label[1:]
                CHANGED = True
            parent = row[parent_column].value
            if CONVERT_TO_LOWERCASE and parent:
                row[parent_column].value = parent[0].lower() + parent[1:]
            if label in label_id_map:
                if not row[0].value or len(row[0].value)==0:
                    row[0].value = label_id_map[label]
                    CHANGED = True

        if CHANGED:
            wb.save(file)





# Do this only when external entities change

# Get the external entities from the file and display in the right format for the imports
if False:
    file = 'imports/external-entities.csv'
    with open (file, 'r') as csvfile:
        csvreader = csv.reader(csvfile)
        idslabels = {}

        for row in csvreader:
            rowdata = row
            id = rowdata[0]
            label = rowdata[1]
            idslabels[id] = label

        for (id,label) in sorted(idslabels.items()):
            #if "CHEBI" in id:
            print(label, ' [', id, ']', sep='', end=';')


if False:

    # External parents/targets of relations:
    # We need to load the information for these from their source ontologies, where they are not included as rows in the spreadsheets.
    # Prepare subsets from several ontologies and merge them

    from ontoutils.robot_wrapper import RobotImportsWrapper

    robotWrapper = RobotImportsWrapper(robotcmd='~/Work/Onto/robot/robot',cleanup=False)
    robotWrapper.processImportsFromExcel(importsFileName='imports/External_Imports.xlsx',
                                            importsOWLURI='http://addictovocab.org/addicto_external.owl',
                                            importsOWLFileName = 'addicto_external.owl',
                                            ontologyName = 'ADDICTO')
    robotWrapper.addAdditionalContent(extraContentTemplate = 'imports/External_Imports_New_Parents.csv',                                         importsOWLURI='http://addictovocab.org/addicto_external.owl',importsOWLFileName = 'addicto_external.owl')

    robotWrapper.addAdditionalContent(extraContentTemplate = 'imports/External_Imports_New_Labels.csv',                                         importsOWLURI='http://addictovocab.org/addicto_external.owl',importsOWLFileName = 'addicto_external.owl')

    robotWrapper.removeProblemMetadata(importsOWLURI='http://addictovocab.org/addicto_external.owl',importsOWLFileName = 'addicto_external.owl', metadataURIFile = "problem-metadata.txt")

    robotWrapper.createOBOFile(importsOWLURI='http://addictovocab.org/addicto_external.owl',importsOWLFileName = 'addicto_external.owl')

# ~/Work/Onto/robot/robot merge --input temp/bcio_upper_level.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates all --output temp/BCIO-slim.owl --prefix "BCIO: http://humanbehaviourchange.org/ontology/BCIO_" --lower-term BCIO:036000 --lower-term BCIO:040000 --lower-term BCIO:034000 --lower-term BCIO:042000 --lower-term BCIO:038000 --lower-term BCIO:043000 --lower-term BCIO:003000 --lower-term BCIO:037000 --lower-term BCIO:041000
#~/Work/Onto/robot/robot merge --input temp/obcs.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates all --output temp/OBCS-slim.owl --lower-term OBCS:0000071 --lower-term OBCS:0000150 --lower-term OBCS:0000035
#~/Work/Onto/robot/robot merge --input temp/chebi.owl extract --method MIREOT --annotate-with-source true --upper-term CHEBI:24431 --intermediates minimal --output temp/CHEBI-CHEM-slim.owl --lower-term CHEBI:2679 --lower-term CHEBI:2972 --lower-term CHEBI:22720 --lower-term CHEBI:3216 --lower-term CHEBI:3219 --lower-term CHEBI:27732 --lower-term CHEBI:67194 --lower-term CHEBI:59999 --lower-term CHEBI:27958 --lower-term CHEBI:38164 --lower-term CHEBI:16236 --lower-term CHEBI:119915 --lower-term CHEBI:16842 --lower-term CHEBI:42797 --lower-term CHEBI:27808 --lower-term CHEBI:6807 --lower-term CHEBI:59331 --lower-term CHEBI:6809 --lower-term CHEBI:7459 --lower-term CHEBI:18723 --lower-term CHEBI:32692 --lower-term CHEBI:67201 --lower-term CHEBI:66964 --lower-term CHEBI:84500 --lower-term CHEBI:35803 --lower-term CHEBI:127342 --lower-term CHEBI:16482 --lower-term CHEBI:17153 --lower-term CHEBI:17245 --lower-term CHEBI:18723 --lower-term CHEBI:36586 --lower-term CHEBI:37249 --lower-term CHEBI:39106 --lower-term CHEBI:4055 --lower-term CHEBI:41607 --lower-term CHEBI:49575 --lower-term CHEBI:5779 --lower-term CHEBI:69478 --lower-term CHEBI:7852 --lower-term CHEBI:34967 --lower-term CHEBI:1391
#~/Work/Onto/robot/robot merge --input temp/chebi.owl extract --method MIREOT --annotate-with-source true --upper-term CHEBI:24432 --intermediates minimal --output temp/CHEBI-BIO-ROLE-slim.owl --lower-term CHEBI:50903 --lower-term CHEBI:50269 --lower-term CHEBI:35617 --lower-term CHEBI:47958 --lower-term CHEBI:48878 --lower-term CHEBI:50903 --lower-term CHEBI:67072 --lower-term CHEBI:73416 --lower-term CHEBI:73417
#~/Work/Onto/robot/robot merge --input temp/chebi.owl extract --method MIREOT --annotate-with-source true --upper-term CHEBI:52217 --intermediates minimal --output temp/CHEBI-DRUG-ROLE-slim.owl --lower-term CHEBI:35469 --lower-term CHEBI:35477 --lower-term CHEBI:35476 --lower-term CHEBI:35474 --lower-term CHEBI:35488 --lower-term CHEBI:35470 --lower-term CHEBI:52217 --lower-term CHEBI:35473 --lower-term CHEBI:23888 --lower-term CHEBI:35471 --lower-term CHEBI:35482 --lower-term CHEBI:51177
#~/Work/Onto/robot/robot merge --input temp/envo.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates all --output temp/ENVO-slim.owl --lower-term ENVO:00002221 --lower-term ENVO:00010483 --lower-term ENVO:00010505 --lower-term ENVO:01000786 --lower-term ENVO:01000838
#~/Work/Onto/robot/robot merge --input temp/sepio.owl extract --method MIREOT --annotate-with-source true --upper-term OBI:0000011 --intermediates all --output temp/SEPIO-slim.owl --lower-term SEPIO:0000004 --lower-term SEPIO:0000125
#~/Work/Onto/robot/robot merge --input temp/po.owl extract --method MIREOT --annotate-with-source true --upper-term PO:0025131 --intermediates minimal --output temp/PO-slim.owl --lower-term PO:0000003
#~/Work/Onto/robot/robot merge --input temp/apollo_sv.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates all --output temp/APOLLO_SV-slim.owl --lower-term APOLLO_SV:00000298
#~/Work/Onto/robot/robot merge --input temp/obi.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates all --output temp/OBI-slim.owl --lower-term OBI:0000011 --lower-term OBI:0000245 --lower-term OBI:0000984 --lower-term OBI:0000423
#~/Work/Onto/robot/robot merge --input temp/omrse.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates all --output temp/OMRSE-slim.owl --lower-term OMRSE:00000062 --lower-term OMRSE:00000102 --lower-term OMRSE:00000114 --lower-term OMRSE:00000106
#~/Work/Onto/robot/robot merge --input temp/uberon.owl extract --method MIREOT --annotate-with-source true --upper-term UBERON:0001062 --intermediates all --output temp/UBERON-ANATOMY-slim.owl --lower-term UBERON:0000467 --lower-term UBERON:0000062 --lower-term UBERON:0001007 --lower-term UBERON:0001873 --lower-term UBERON:0004535 --lower-term UBERON:0006314
#~/Work/Onto/robot/robot merge --input temp/uberon.owl extract --method MIREOT --annotate-with-source true --upper-term UBERON:0000000 --intermediates all --output temp/UBERON-PROCESS-slim.owl --lower-term UBERON:0035943 --lower-term UBERON:035944 --lower-term UBERON:0000071
#~/Work/Onto/robot/robot merge --input temp/doid.owl extract --method MIREOT --annotate-with-source true --upper-term DOID:4 --intermediates minimal --output temp/DOID-slim.owl --lower-term DOID:0050668 --lower-term DOID:0060903 --lower-term DOID:10935 --lower-term DOID:10937 --lower-term DOID:12995 --lower-term DOID:1470 --lower-term DOID:1510 --lower-term DOID:162 --lower-term DOID:2030 --lower-term DOID:2468 --lower-term DOID:3312 --lower-term DOID:399 --lower-term DOID:526
#~/Work/Onto/robot/robot merge --input temp/core.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates minimal --output temp/RO-CORE-slim.owl --lower-term BFO:0000051 --lower-term RO:0000053 --lower-term RO:0000052 --lower-term RO:0001025 --lower-term RO:0000056 --lower-term RO:0000057 --lower-term RO:0000091 --lower-term RO:0001000 --lower-term RO:0000087 --lower-term BFO:0000050 --lower-term RO:0000086
#~/Work/Onto/robot/robot merge --input temp/ro.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates minimal --output temp/RO-slim.owl --lower-term RO:0002234 --lower-term BFO:0000055 --lower-term BFO:0000054 --lower-term RO:0002353 --lower-term RO:0001019 --lower-term RO:0002577
#~/Work/Onto/robot/robot merge --input temp/bfo.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates minimal --output temp/BFO-DEV-slim.owl --lower-term BFO:0000117 --lower-term BFO:0000119
#~/Work/Onto/robot/robot merge --input temp/bfo.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates all --output temp/BFO-slim.owl --lower-term BFO:0000015 --lower-term BFO:0000027 --lower-term BFO:0000023 --lower-term BFO:0000016 --lower-term BFO:0000030 --lower-term BFO:0000144 --lower-term BFO:0000029 --lower-term BFO:0000003 --lower-term BFO:0000035 --lower-term BFO:0000008 --lower-term ﻿BFO:0000024 --lower-term BFO:0000141
#~/Work/Onto/robot/robot merge --input temp/iao.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates all --output temp/IAO-slim.owl --lower-term IAO:0000136 --lower-term IAO:0000027 --lower-term IAO:0000088 --lower-term IAO:0000104 --lower-term IAO:0000178 --lower-term IAO:0000310 --lower-term IAO:0000007 --lower-term IAO:0000033
#~/Work/Onto/robot/robot merge --input temp/mf.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates all --output temp/MF-slim.owl --lower-term MF:0000016 --lower-term MF:0000033 --lower-term MF:0000031 --lower-term MF:0000020
#~/Work/Onto/robot/robot merge --input temp/mfoem.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates all --output temp/MFOEM-slim.owl --lower-term MFOEM:000006 --lower-term MFOEM:000005 --lower-term MFOEM:000001
#~/Work/Onto/robot/robot merge --input temp/pato.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000002 --intermediates all --output temp/PATO-slim.owl --lower-term PATO:0001018 --lower-term PATO:0000033
#~/Work/Onto/robot/robot merge --input temp/ogms.owl extract --method MIREOT --annotate-with-source true --upper-term OGMS:0000045 --intermediates all --output temp/OGMS-slim.owl --lower-term OGMS:0000102 --lower-term OGMS:0000031
#~/Work/Onto/robot/robot merge --input temp/chmo.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000015 --intermediates all --output temp/CHMO-slim.owl --lower-term CHMO:0001000 --lower-term CHMO:0001004
#~/Work/Onto/robot/robot merge --input temp/ero.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000015 --intermediates minimal --output temp/ERO-slim.owl --lower-term ERO:0001108
#~/Work/Onto/robot/robot merge --input temp/oae.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates minimal --output temp/OAE-slim.owl --lower-term OAE:0001000 --lower-term OAE:0000001 --lower-term OAE:0003232
#~/Work/Onto/robot/robot merge --input temp/gsso.owl extract --method MIREOT --annotate-with-source true --upper-term BFO:0000001 --intermediates none --output temp/GSSO-slim.owl --lower-term GSSO:000130 --lower-term GSSO:000137 --lower-term GSSO:000369 --lower-term GSSO:000370 --lower-term GSSO:000376 --lower-term GSSO:000379 --lower-term GSSO:000395 --lower-term GSSO:000498 --lower-term GSSO:000529 --lower-term GSSO:000924 --lower-term GSSO:001590 --lower-term GSSO:001595 --lower-term GSSO:001596 --lower-term GSSO:001802 --lower-term GSSO:002821 --lower-term GSSO:002957 --lower-term GSSO:002961 --lower-term GSSO:002962 --lower-term GSSO:003501 --lower-term GSSO:004516 --lower-term GSSO:004615 --lower-term GSSO:005301 --lower-term GSSO:005379 --lower-term GSSO:007328 --lower-term GSSO:009381 --prefix "GSSO: http://purl.obolibrary.org/obo/GSSO_"
#~/Work/Onto/robot/robot merge --input temp/BCIO-slim.owl --input temp/OBCS-slim.owl --input temp/CHEBI-CHEM-slim.owl --input temp/CHEBI-BIO-ROLE-slim.owl --input temp/CHEBI-DRUG-ROLE-slim.owl --input temp/ENVO-slim.owl --input temp/SEPIO-slim.owl --input temp/PO-slim.owl --input temp/APOLLO_SV-slim.owl --input temp/OBI-slim.owl --input temp/OMRSE-slim.owl --input temp/UBERON-ANATOMY-slim.owl --input temp/UBERON-PROCESS-slim.owl --input temp/DOID-slim.owl --input temp/RO-CORE-slim.owl --input temp/RO-slim.owl --input temp/BFO-DEV-slim.owl --input temp/BFO-slim.owl --input temp/IAO-slim.owl --input temp/MF-slim.owl --input temp/MFOEM-slim.owl --input temp/PATO-slim.owl --input temp/OGMS-slim.owl --input temp/CHMO-slim.owl --input temp/ERO-slim.owl --input temp/OAE-slim.owl --input temp/GSSO-slim.owl annotate --ontology-iri http://addictovocab.org/addicto_external.owl --version-iri http://addictovocab.org/addicto_external.owl --annotation rdfs:comment  "This file contains externally imported content for the ADDICTO. It was prepared using ROBOT and a custom script from a spreadsheet of imported terms." --output addicto_external.owl
#About to execute Robot command:  ~/Work/Onto/robot/robot template --template imports/External_Imports_New_Parents.csv --ontology-iri http://addictovocab.org/addicto_external.owl --output addicto_external-temp.owl
#About to execute Robot command:  ~/Work/Onto/robot/robot merge --input addicto_external.owl --input addicto_external-temp.owl --output addicto_external.owl
#About to execute Robot command:  ~/Work/Onto/robot/robot template --template imports/External_Imports_New_Labels.csv --ontology-iri http://addictovocab.org/addicto_external.owl --output addicto_external-temp.owl
#About to execute Robot command:  ~/Work/Onto/robot/robot merge --input addicto_external.owl --input addicto_external-temp.owl --output addicto_external.owl
#About to execute Robot command:  ~/Work/Onto/robot/robot remove --input addicto_external.owl --term-file problem-metadata.txt --axioms annotation --output addicto_external.owl
#About to execute Robot command:  ~/Work/Onto/robot/robot merge --input addicto_external.owl convert --output addicto_external.obo --check false


# load the dictionary for prefix to URI mapping

dict_file = 'scripts/prefix_to_uri_dictionary.csv'
prefix_dict = {}
reader = csv.DictReader(open(dict_file, 'r'))
for line in reader:
    prefix_dict[line['PREFIX']] = line['URI_PREFIX']

path = 'outputs'


# Load the output files

addicto_files = []

for root, dirs_list, files_list in os.walk(path):
    for file_name in files_list:
        full_file_name = os.path.join(root, file_name)
        addicto_files.append(full_file_name)



entries = {}

# All the external content + hierarchy is in the file addicto_external.obo
# Must be parsed and sent to AOVocab.

externalonto = pronto.Ontology("addicto_external.obo")
for term in externalonto.terms():
    label_id_map[term.name] = term.id

# get the "ready" input data, indexed by ID
for filename in addicto_files:
    with open(filename, 'r') as csvfile:
        csvreader = csv.reader(csvfile)
        header = next(csvreader)

        for row in csvreader:
            rowdata = row
            id = rowdata[0]
            entries[id] = (header,rowdata)


# Merge into a single spreadsheet for creating an input file for ROBOT to create an OWL file (entries built in submit to addicto vocab file)
# First get merger of used headers
all_headers = []
headers_to_exclude = ['AO sub-ontology', 'Upper level', 'Fuzzy set', 'E-CigO', 'Proposer', 'Curation status', 'Why fuzzy', 'Cross reference', 'Type']
rel_mappings = {'has part':'[BFO:0000051]',
                'has role': '[RO:0000087]',
                'Derives from':'[RO:0001000]',
                'is about' :'[IAO:0000136]',
                'contains' : '[RO:0001019]'
                }
for id in entries:
    (header,rowdata) = entries[id]
    for item in header:
        if item not in all_headers and item not in headers_to_exclude:
            all_headers.append(item)

all_row_data = []

for id in entries:
    (header,rowdata) = entries[id]
    new_rowdata = []

    # exclude external entities
    if "ADDICTO" not in rowdata[0]:
        continue

    # exclude entities with nonexistent parents
    parentStr = rowdata[header.index('Parent')]
    #print("Got parentstr: ",parentStr)
    parentVals = parentStr.split(";")
    parents = []
    for p in parentVals:
        m = re.search(r"\[([A-Za-z0-9:]+)\]", p)
        if m:
            p = m.group(1)
            print ("Found match:", m, "=>", p)
        if re.fullmatch("[A-Za-z]*:(\d)*",p):
            #print("Got ID-form parent: ",p, "replacing with label", getLabelForID(p))
            p = getLabelForID(p)
            parents.append(p)
        else:
            try:
                id = getIdForLabel(p)
            except ValueError:
                print("No ID found for ",p)
                continue # Not found
            else:
                parents.append(getCorrectFormForLabel(p)) # it has an id, but, we want the label
    if len(parents)>0:
        parent = ";".join(parents)
    else:
        #print ("No usable parents in",parentStr)
        continue # SKIP THIS ENTRY

    for h in all_headers:
        if h in header:
            if h == 'Parent':
                new_rowdata.append(parent)
            else:
                i = header.index(h)
                val = rowdata[i]
                new_rowdata.append(val)
        else:
            new_rowdata.append('')
    all_row_data.append(new_rowdata)


inputFileName  = 'AddictO-own-merged.xlsx'


from openpyxl import Workbook
wb = Workbook()
sheet = wb.active
# Write header
for i, v in enumerate(all_headers):
    for mp in rel_mappings.keys():
        if mp in v:
            v = v + " "+rel_mappings[mp]
    sheet.cell(row=1,column=i+1).value = v
# Write values
for i, line in enumerate(all_row_data):
    for k, val in enumerate(line):
        sheet.cell(row=i+2, column=k+1).value = val
wb.save(inputFileName)


#from importlib import reload
#reload(ontoutils.robot_wrapper)

from ontoutils.robot_wrapper import RobotTemplateWrapper
robotWrapper = RobotTemplateWrapper(robotcmd='~/Work/Onto/robot/robot')

csvFileName = robotWrapper.processClassInfoFromExcel(inputFileName)


## EXECUTE THE ROBOT COMMAND AS A SUB-PROCESS

owlFileName = "addicto.owl"
IRI_PREFIX = 'http://addictovocab.org/'
ID_PREFIX = '\"ADDICTO: '+IRI_PREFIX+'ADDICTO_\"'
ONTOLOGY_IRI = IRI_PREFIX+owlFileName
dependency='addicto_external.owl'


robotWrapper.createOntologyFromTemplateFile(csvFileName, dependency, IRI_PREFIX, ID_PREFIX, ONTOLOGY_IRI,owlFileName)

from ontoutils.robot_wrapper import RobotImportsWrapper
robotWrapper = RobotImportsWrapper(robotcmd='~/Work/Onto/robot/robot',cleanup=False)
robotWrapper.createOBOFile(importsOWLURI='http://addictovocab.org/addicto.owl',importsOWLFileName = 'addicto.owl')

# Annotate with version information
# robot annotate --input addicto.owl --annotation rdfs:comment "The Addiction Ontology (AddictO) is an ontology being developed all aspects of addiction research." --annotation owl:versionInfo "2022-02-16" --ontology-iri "http://addictovocab.org/addicto.owl" --version-iri "http://addictovocab.org/addicto.owl/2022-02-16" --output addicto.owl
#
# Build the merged file for onto-edit
# robot merge --input addicto.owl convert --output addicto-merged.owx

